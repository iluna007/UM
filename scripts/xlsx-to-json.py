#!/usr/bin/env python3
"""
Export each sheet of Airbnb.xlsx, Google Review.xlsx, and Noticias.xlsx to JSON.
Each workbook becomes one JSON file: { "source": "filename", "sheets": { "SheetName": [rows] } }.
Run from project root: python scripts/xlsx-to-json.py
Output: src/data/airbnb.json, src/data/google-review.json, src/data/noticias.json
"""

import json
import re
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    raise


def to_serializable(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.isoformat() if hasattr(val, "isoformat") else str(val)
    if isinstance(val, float) and val == int(val):
        return int(val)
    return val

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "src" / "data"
OUT_NAMES = {
    "Airbnb.xlsx": "airbnb.json",
    "Google Review.xlsx": "google-review.json",
    "Noticias.xlsx": "noticias.json",
}


def slug(name):
    """Turn sheet name into a safe key."""
    return re.sub(r"[^\w\-]", "_", (name or "Sheet").strip())[:80]


def row_to_dict(headers, row):
    out = {}
    for i, h in enumerate(headers):
        val = row[i] if i < len(row) else None
        val = to_serializable(val)
        key = (h or f"col_{i}").strip() or f"col_{i}"
        out[key] = val
    return out


def sheet_to_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    return [row_to_dict(headers, list(r)) for r in rows[1:]]


def xlsx_to_json(path: Path, out_name: str) -> None:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = sheet_to_rows(ws)
        sheets[name] = rows
    wb.close()
    data = {"source": path.name, "sheetNames": wb.sheetnames, "sheets": sheets}
    out_path = DATA_DIR / out_name
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"Wrote {out_path} ({len(sheets)} sheets)")


def main():
    for filename, out_name in OUT_NAMES.items():
        path = DATA_DIR / filename
        if not path.exists():
            print(f"Skip (not found): {path}")
            continue
        xlsx_to_json(path, out_name)
    print("Done.")


if __name__ == "__main__":
    main()
